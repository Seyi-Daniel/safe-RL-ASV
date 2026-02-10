#!/usr/bin/env python3
"""Convert a sensitivity_summary.json report into table-friendly CSV/XLSX files."""
from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path
from typing import Any


TABLE_COLUMNS = [
    "Step",
    "Explanation",
    "Output Channel",
    "Feature",
    "Feature Attribution Score",
    "Original Output",
    "Perturbation",
    "New Output",
    "Difference",
    "Original Label",
    "Label Change",
    "New Label",
]


def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--input",
        type=Path,
        required=True,
        help="Path to a sensitivity_summary.json file.",
    )
    parser.add_argument(
        "--csv-out",
        type=Path,
        default=None,
        help="Output CSV path (defaults to <input_stem>_table.csv).",
    )
    parser.add_argument(
        "--xlsx-out",
        type=Path,
        default=None,
        help="Output Excel path (defaults to <input_stem>_table.xlsx).",
    )
    parser.add_argument(
        "--skip-xlsx",
        action="store_true",
        help="Only write CSV and skip XLSX generation.",
    )
    return parser


def _load_json(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as fh:
        payload = json.load(fh)
    if not isinstance(payload, dict) or "steps" not in payload:
        raise ValueError(
            f"Input JSON must be a scenario sensitivity summary containing a 'steps' key: {path}"
        )
    return payload


def _format_float(value: float) -> str:
    return f"{value:.6f}"


def _flatten_rows(payload: dict[str, Any]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for step_payload in payload.get("steps", []):
        step = int(step_payload["step"])
        baseline_outputs = step_payload["baseline"]["outputs"]

        for explanation_name in ("lime", "shap"):
            explainer_payload = step_payload.get("explainers", {}).get(explanation_name, {})
            if not explainer_payload:
                continue

            for output_name in ("rudder", "throttle"):
                output_payload = explainer_payload.get(output_name, {})
                for feature_payload in output_payload.get("sensitivity", []):
                    feature_name = feature_payload["feature"]
                    feature_attrib = float(feature_payload["attribution"])

                    for perturbation in feature_payload.get("perturbations", []):
                        direction = perturbation["direction"]
                        percent = float(perturbation["percent"])
                        outputs = perturbation["outputs"]
                        deltas = perturbation["deltas"]

                        if output_name == "rudder":
                            original_output = float(baseline_outputs["rudder"])
                            new_output = float(outputs["rudder"])
                            diff = float(deltas["rudder_delta"])
                            original_label = str(baseline_outputs["helm_label"])
                            label_changed = bool(deltas["helm_label_changed"])
                            new_label = str(outputs["helm_label"])
                        else:
                            original_output = float(baseline_outputs["throttle_raw"])
                            new_output = float(outputs["throttle_raw"])
                            diff = float(deltas["throttle_raw_delta"])
                            original_label = str(baseline_outputs["throttle_label"])
                            label_changed = bool(deltas["throttle_label_changed"])
                            new_label = str(outputs["throttle_label"])

                        rows.append(
                            {
                                "Step": str(step),
                                "Explanation": explanation_name,
                                "Output Channel": output_name,
                                "Feature": feature_name,
                                "Feature Attribution Score": _format_float(feature_attrib),
                                "Original Output": _format_float(original_output),
                                "Perturbation": f"{direction} {percent:g}%",
                                "New Output": _format_float(new_output),
                                "Difference": _format_float(diff),
                                "Original Label": original_label,
                                "Label Change": "Yes" if label_changed else "No",
                                "New Label": new_label,
                            }
                        )
    return rows


def _write_csv(rows: list[dict[str, str]], csv_out: Path) -> None:
    csv_out.parent.mkdir(parents=True, exist_ok=True)
    with csv_out.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=TABLE_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)


def _merge_repeated_cells(worksheet, min_row: int, max_row: int, col_idx: int) -> None:
    if max_row <= min_row:
        return
    start = min_row
    prev_value = worksheet.cell(row=min_row, column=col_idx).value
    for row in range(min_row + 1, max_row + 1):
        value = worksheet.cell(row=row, column=col_idx).value
        if value != prev_value:
            if row - 1 > start:
                worksheet.merge_cells(start_row=start, end_row=row - 1, start_column=col_idx, end_column=col_idx)
            start = row
            prev_value = value
    if max_row > start:
        worksheet.merge_cells(start_row=start, end_row=max_row, start_column=col_idx, end_column=col_idx)


def _write_xlsx(rows: list[dict[str, str]], xlsx_out: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover - optional dependency
        raise RuntimeError(
            "openpyxl is required to write XLSX output. Install it with: pip install openpyxl"
        ) from exc

    xlsx_out.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sensitivity"

    worksheet.append(TABLE_COLUMNS)
    for row in rows:
        worksheet.append([row[col] for col in TABLE_COLUMNS])

    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    total_rows = len(rows) + 1
    total_cols = len(TABLE_COLUMNS)

    for col in range(1, total_cols + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    for row in worksheet.iter_rows(min_row=1, max_row=total_rows, min_col=1, max_col=total_cols):
        for cell in row:
            cell.border = border

    for col_idx in (1, 2, 3, 4):
        _merge_repeated_cells(worksheet, min_row=2, max_row=total_rows, col_idx=col_idx)

    for row in worksheet.iter_rows(min_row=2, max_row=total_rows, min_col=1, max_col=4):
        for cell in row:
            cell.alignment = center

    for col_idx in range(1, total_cols + 1):
        max_len = 0
        for row_idx in range(1, total_rows + 1):
            value = worksheet.cell(row=row_idx, column=col_idx).value
            value_len = len(str(value)) if value is not None else 0
            max_len = max(max_len, value_len)
        worksheet.column_dimensions[get_column_letter(col_idx)].width = min(40, max(12, max_len + 2))

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    workbook.save(xlsx_out)


def main() -> None:
    parser = _build_parser()
    args = parser.parse_args()

    input_path = args.input
    payload = _load_json(input_path)
    rows = _flatten_rows(payload)

    csv_out = args.csv_out or input_path.with_name(f"{input_path.stem}_table.csv")
    _write_csv(rows, csv_out)
    print(f"Wrote CSV table: {csv_out}")

    if args.skip_xlsx:
        return

    xlsx_out = args.xlsx_out or input_path.with_name(f"{input_path.stem}_table.xlsx")
    _write_xlsx(rows, xlsx_out)
    print(f"Wrote XLSX table: {xlsx_out}")


if __name__ == "__main__":
    main()
